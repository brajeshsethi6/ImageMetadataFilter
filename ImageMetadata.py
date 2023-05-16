from PIL import Image, ExifTags
import os
import openpyxl
from datetime import datetime
import shutil

# Set the directory path
directory = r"F:\Projects\ImageMetadataFilter"
# Get all the files in the directory
files = os.listdir(directory)
# Filter the files to include only JPG files
#jpg_files = [file for file in files if file.endswith(".jpg",".jpeg")]
jpg_files = [file for file in files if file.endswith((".jpg", ".jpeg"))]

print("Name of the file are",jpg_files)

# Load the existing Excel file or create a new one if it doesn't exist
excel_file = "metadata.xlsx"
if os.path.exists(os.path.join(directory, excel_file)):
    workbook = openpyxl.load_workbook(os.path.join(directory, excel_file))
else:
    workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Define the date format for the worksheet
date_format = "dd-mm-yyyy"

# If the worksheet is empty, write the header row
if worksheet.max_row == 1:
    worksheet.cell(row=1, column=1, value="File Name")
    worksheet.cell(row=1, column=2, value="DateTime")
    worksheet.cell(row=1, column=3, value="Validation") #Less than 5 days
    worksheet.cell(row=1, column=4, value="Date of check in")
    worksheet.cell(row=1, column=5, value="Name")

# Write the metadata to the worksheet for each image
row_num = worksheet.max_row + 1
for img_file in jpg_files:
    # Open the image file
    print(img_file)
    img = Image.open(os.path.join(directory, img_file))

    # Get the metadata
    metadata = img._getexif()
#in some case metadata is coming as None ?? 
    if metadata:
        img.close()
        for tag_id in metadata:
            tag_name = ExifTags.TAGS.get(tag_id, tag_id)
            if tag_name == "DateTime":
                tag_value = metadata.get(tag_id)
                try:
                    # Extract the date value from the metadata
                    date_value = datetime.strptime(tag_value, "%Y:%m:%d %H:%M:%S").date()
                    if (date_value):
                        # Write the file name, date value, and whether the date is less than 5 days from today to the worksheet
                        worksheet.cell(row=row_num, column=1, value=img_file)
                        worksheet.cell(row=row_num, column=2, value=date_value).number_format = date_format
                        
                        if (datetime.today().date() - date_value).days <= 5:
                            worksheet.cell(row=row_num, column=3, value="Successful")
                            #This code should be below the Successful 
                            try:
                                worksheet.cell(row=row_num, column=5, value=img_file.split(".")[0].split("-")[1].strip())
                            except:
                                worksheet.cell(row=row_num, column=5, value="NA")
                        else:
                            worksheet.cell(row=row_num, column=3, value="Unsuccessful")
                            #worksheet.cell(row=row_num, column=5, value="Unsuccessful")
                            try:
                                worksheet.cell(row=row_num, column=5, value=img_file.split(".")[0].split("-")[1].strip())
                            except:
                                worksheet.cell(row=row_num, column=5, value="NA")
                        worksheet.cell(row=row_num, column=4, value=datetime.today().date()).number_format = date_format

                        ##########
                        # #This code should be below the Successful 
                        # try:
                        #     worksheet.cell(row=row_num, column=5, value=img_file.split(".")[0].split("-")[1].strip())
                        # except:
                        #     worksheet.cell(row=row_num, column=5, value="NA")
                        ##########

                        row_num += 1
  
                except ValueError:
                    # Skip any values that cannot be converted to a date
                    pass
    else:
        print("No Metadata")
# Save the Excel file
try:
    workbook.save(os.path.join(directory, excel_file))
    print(f"Metadata saved to {excel_file}")
except:
    print("Close the Excel file first")

# Set the source and destination directory paths
src_directory = r"F:\Projects\ImageMetadataFilter"
dst_directory = r"F:\Projects\ImageMetadataFilter\ProcessedImages"

# Create the destination directory if it doesn't exist
if not os.path.exists(dst_directory):
    os.makedirs(dst_directory)
img.close()
# Move each image file from the source directory to the destination directory
for img_file in jpg_files:
    src_path = os.path.join(src_directory, img_file)
    dst_path = os.path.join(dst_directory, img_file)
    shutil.move(src_path, dst_path)

if (len(jpg_files) == 0):
    print("ImageMetadata Folder is Empty")
else:
    print(f"All image files moved to {dst_directory}")
########################################################
#if the file exced then 30 delete the olderst file

store_files = os.listdir(dst_directory)

if len(store_files) >= 31:
    store_files_sorted = sorted(store_files, key=lambda f: os.path.getmtime(os.path.join(dst_directory, f)))

    while len(store_files_sorted) >= 30:
        # Remove the oldest file
        os.remove(os.path.join(dst_directory, store_files_sorted[0]))
        print("Removed oldest file:", store_files_sorted[0])

        # Update the sorted list of files
        store_files_sorted = sorted(store_files, key=lambda f: os.path.getmtime(os.path.join(dst_directory, f)))
else:
    print("Number of files is below 30. No files will be removed.")